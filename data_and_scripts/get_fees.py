import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import requests
import os
import openpyxl
import matplotlib.pyplot as plt
import math

def fetch_and_analyze_blockchain_data(start_date="2023-02-08", end_date="2025-02-08"):
    """
    Fetch blockchain data and analyze it to understand the pattern
    
    Args:
        start_date (str): Start date in 'YYYY-MM-DD' format
        end_date (str): End date in 'YYYY-MM-DD' format
    
    Returns:
        pandas.DataFrame: Raw blockchain data
    """
    # Convert dates to datetime objects
    start_dt = pd.to_datetime(start_date)
    end_dt = pd.to_datetime(end_date)
    
    # Calculate timespan
    days_diff = (end_dt - start_dt).days
    if days_diff <= 30:
        timespan = f"{days_diff}days"
    elif days_diff <= 90:
        timespan = f"{days_diff//7 + 1}weeks"
    elif days_diff <= 730:
        timespan = f"{days_diff//30 + 1}months"
    else:
        timespan = f"{days_diff//365 + 1}years"
    
    print(f"Fetching data from {start_date} to {end_date} (timespan: {timespan})")
    
    # Fetch transaction fees data
    url = f'https://api.blockchain.info/charts/transaction-fees?timespan={timespan}&format=json&sampled=true&rollingAverage=3hours'
    print(f"Fetching {url}")
    
    response = requests.get(url)
    if response.status_code != 200:
        print(f"Error fetching transaction fees: {response.status_code}")
        return pd.DataFrame()
    
    data = response.json()
    if 'values' not in data or not data['values']:
        print("No transaction fees data found")
        return pd.DataFrame()
    
    df = pd.DataFrame(data['values'])
    df.columns = ['timestamp', 'transaction_fees_btc']
    df['timestamp'] = pd.to_datetime(df['timestamp'], unit='s')
    
    # Fetch market price data
    url = f'https://api.blockchain.info/charts/market-price?timespan={timespan}&format=json&sampled=true&rollingAverage=3hours'
    print(f"Fetching {url}")
    
    response = requests.get(url)
    if response.status_code != 200:
        print(f"Error fetching market price: {response.status_code}")
        return df  # Return just the fees without price
    
    data = response.json()
    if 'values' in data and data['values']:
        price_df = pd.DataFrame(data['values'])
        price_df.columns = ['timestamp', 'btc_market_price_usd']
        price_df['timestamp'] = pd.to_datetime(price_df['timestamp'], unit='s')
        
        # Merge price data with fees data
        df = pd.merge(df, price_df, on='timestamp', how='outer')
        df.sort_values('timestamp', inplace=True)
        
        # Fill missing values
        df = df.ffill().bfill()
    
    # Calculate transaction fees in USD
    if 'btc_market_price_usd' in df.columns:
        df['transaction_fees_usd'] = df['transaction_fees_btc'] * df['btc_market_price_usd']
    
    return df

def analyze_original_fees(excel_file):
    """
    Analyze the original transaction fee data from the Excel file
    
    Args:
        excel_file (str): Path to the Excel file
        
    Returns:
        pandas.DataFrame: DataFrame with the original transaction fee data
    """
    print(f"Analyzing original transaction fee data from {excel_file}")
    
    # Read the Excel file
    df = pd.read_excel(excel_file)
    
    # Find the transaction fee columns
    btc_fee_col = None
    usd_fee_col = None
    
    for col in df.columns:
        col_lower = col.lower()
        if 'transaction' in col_lower and 'fee' in col_lower and 'btc' in col_lower:
            btc_fee_col = col
        elif 'transaction' in col_lower and 'fee' in col_lower and 'usd' in col_lower:
            usd_fee_col = col
    
    if not btc_fee_col:
        print("Error: Could not find BTC transaction fee column")
        return None
    
    # Filter to only the columns we need
    result = df[['timestamp', btc_fee_col]].copy()
    if usd_fee_col:
        result[usd_fee_col] = df[usd_fee_col]
    
    # Clean up the data
    result.dropna(subset=['timestamp'], inplace=True)
    result.rename(columns={btc_fee_col: 'original_btc_fees'}, inplace=True)
    if usd_fee_col:
        result.rename(columns={usd_fee_col: 'original_usd_fees'}, inplace=True)
    
    # Sort by timestamp
    result['timestamp'] = pd.to_datetime(result['timestamp'])
    result.sort_values('timestamp', inplace=True)
    
    # Calculate statistics
    print("Original transaction fee statistics:")
    print(f"Number of records: {len(result)}")
    print(f"Min BTC fee: {result['original_btc_fees'].min()}")
    print(f"Max BTC fee: {result['original_btc_fees'].max()}")
    print(f"Mean BTC fee: {result['original_btc_fees'].mean()}")
    print(f"Median BTC fee: {result['original_btc_fees'].median()}")
    
    # Check for jumps in the data
    fees = result['original_btc_fees'].values
    jumps = []
    for i in range(1, len(fees)):
        if fees[i-1] > 0 and fees[i] > 0:
            ratio = fees[i] / fees[i-1]
            if ratio > 2 or ratio < 0.5:
                jumps.append({
                    'index': i,
                    'timestamp': result['timestamp'].iloc[i],
                    'prev': fees[i-1],
                    'current': fees[i],
                    'ratio': ratio
                })
    
    if jumps:
        print(f"\nFound {len(jumps)} significant jumps in the original data")
        for jump in jumps[:5]:
            print(f"  {jump['timestamp']}: {jump['prev']} → {jump['current']} ({jump['ratio']:.2f}x)")
    else:
        print("\nNo significant jumps found in the original data")
    
    return result

def plot_fee_comparison(original_df, blockchain_df, output_file='fee_comparison.png'):
    """
    Plot a comparison of the original fees and the blockchain fees
    
    Args:
        original_df (pandas.DataFrame): DataFrame with original fee data
        blockchain_df (pandas.DataFrame): DataFrame with blockchain fee data
        output_file (str): Path to save the plot
    """
    plt.figure(figsize=(12, 8))
    
    # Plot original fees
    plt.subplot(2, 1, 1)
    plt.plot(original_df['timestamp'], original_df['original_btc_fees'], label='Original BTC Fees')
    plt.title('Original Transaction Fees')
    plt.ylabel('BTC')
    plt.grid(True)
    plt.legend()
    
    # Plot blockchain fees
    plt.subplot(2, 1, 2)
    plt.plot(blockchain_df['timestamp'], blockchain_df['transaction_fees_btc'], label='Blockchain API BTC Fees')
    plt.title('Blockchain API Transaction Fees')
    plt.xlabel('Date')
    plt.ylabel('BTC')
    plt.grid(True)
    plt.legend()
    
    plt.tight_layout()
    plt.savefig(output_file)
    print(f"Fee comparison plot saved to {output_file}")

def manually_scale_blockchain_data(original_df, blockchain_df):
    """
    Manually scale the blockchain data to match the original data's pattern
    
    Args:
        original_df (pandas.DataFrame): DataFrame with original fee data
        blockchain_df (pandas.DataFrame): DataFrame with blockchain fee data
        
    Returns:
        pandas.DataFrame: Scaled blockchain data
    """
    # Get scaling factor from the first day of data that's present in both datasets
    original_start_date = original_df['timestamp'].min().strftime('%Y-%m-%d')
    blockchain_df_day1 = blockchain_df[blockchain_df['timestamp'].dt.strftime('%Y-%m-%d') == original_start_date]
    original_df_day1 = original_df[original_df['timestamp'].dt.strftime('%Y-%m-%d') == original_start_date]
    
    if not blockchain_df_day1.empty and not original_df_day1.empty:
        blockchain_day1_avg = blockchain_df_day1['transaction_fees_btc'].mean()
        original_day1_avg = original_df_day1['original_btc_fees'].mean()
        
        if blockchain_day1_avg > 0 and original_day1_avg > 0:
            scaling_factor = original_day1_avg / blockchain_day1_avg
            print(f"Calculated scaling factor from first day: {scaling_factor:.6f}")
            print(f"  Original first day average: {original_day1_avg}")
            print(f"  Blockchain first day average: {blockchain_day1_avg}")
        else:
            # Default scaling if we can't calculate from day 1
            scaling_factor = 0.015  # A typical ratio for per-block vs total network fees
            print(f"Using default scaling factor: {scaling_factor}")
    else:
        # Default scaling if we can't find matching days
        scaling_factor = 0.015
        print(f"Using default scaling factor: {scaling_factor}")
    
    # Apply scaling
    result = blockchain_df.copy()
    result['transaction_fees_btc'] = result['transaction_fees_btc'] * scaling_factor
    
    # Recalculate USD fees if available
    if 'btc_market_price_usd' in result.columns and 'transaction_fees_btc' in result.columns:
        result['transaction_fees_usd'] = result['transaction_fees_btc'] * result['btc_market_price_usd']
    
    # Verify the scaling worked
    print("\nAfter scaling:")
    print(f"Min BTC fee: {result['transaction_fees_btc'].min()}")
    print(f"Max BTC fee: {result['transaction_fees_btc'].max()}")
    print(f"Mean BTC fee: {result['transaction_fees_btc'].mean()}")
    print(f"Median BTC fee: {result['transaction_fees_btc'].median()}")
    
    return result

def interpolate_to_3_hour_intervals(df):
    """
    Ensure the data has 3-hour intervals through interpolation
    
    Args:
        df (pandas.DataFrame): DataFrame to interpolate
        
    Returns:
        pandas.DataFrame: DataFrame with 3-hour intervals
    """
    # Set timestamp as index
    df_indexed = df.set_index('timestamp')
    
    # Create a new index with 3-hour intervals
    start_time = df_indexed.index.min()
    end_time = df_indexed.index.max()
    new_index = pd.date_range(start=start_time, end=end_time, freq='3h')
    
    # Reindex and interpolate
    result = df_indexed.reindex(new_index)
    result = result.interpolate(method='time')
    
    # Reset index
    result = result.reset_index().rename(columns={'index': 'timestamp'})
    
    return result

def update_excel_with_scaled_fees(excel_file, scaled_fees_df):
    """
    Update the Excel file with the scaled transaction fee data
    
    Args:
        excel_file (str): Path to the Excel file
        scaled_fees_df (pandas.DataFrame): DataFrame with scaled fee data
    """
    print(f"Updating Excel file {excel_file} with scaled transaction fee data")
    
    # Create a backup of the original file
    backup_file = f"{excel_file.rsplit('.', 1)[0]}_backup_manual.xlsx"
    import shutil
    shutil.copy2(excel_file, backup_file)
    print(f"Created backup of original file: {backup_file}")
    
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    # Find the column indices for transaction fee columns
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_row_list = list(header_row)
    
    try:
        timestamp_col_idx = header_row_list.index("timestamp") + 1  # 1-based indexing in openpyxl
        btc_fees_col_idx = header_row_list.index("transaction fees btc") + 1
        usd_fees_col_idx = header_row_list.index("Average transaction fees usd (3-hour interval)") + 1
        print(f"Found columns: timestamp={timestamp_col_idx}, BTC fees={btc_fees_col_idx}, USD fees={usd_fees_col_idx}")
    except ValueError as e:
        print(f"Error finding column headers: {e}")
        print("Available columns:", header_row_list)
        return
    
    # Create a timestamp to data lookup
    scaled_fees_df['timestamp_str'] = scaled_fees_df['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
    data_lookup = {}
    for _, row in scaled_fees_df.iterrows():
        timestamp_str = row['timestamp_str']
        data_lookup[timestamp_str] = {
            'btc_fees': row['transaction_fees_btc'],
            'usd_fees': row['transaction_fees_usd'] if 'transaction_fees_usd' in row else None
        }
    
    # Update the values in the Excel file
    updates_btc = 0
    updates_usd = 0
    rows_processed = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from row 2 (skip header)
        timestamp_cell = row[timestamp_col_idx - 1]  # Convert to 0-based indexing
        
        # Skip rows without timestamps
        if timestamp_cell.value is None:
            continue
        
        # Format timestamp to match lookup key
        if isinstance(timestamp_cell.value, datetime):
            timestamp_str = timestamp_cell.value.strftime('%Y-%m-%d %H:%M:%S')
        else:
            timestamp_str = str(timestamp_cell.value)
        
        rows_processed += 1
        
        # Update cells if timestamp exists in data
        if timestamp_str in data_lookup:
            # Update BTC fees cell
            btc_fees_cell = row[btc_fees_col_idx - 1]
            new_btc_value = data_lookup[timestamp_str]['btc_fees']
            
            if not math.isnan(new_btc_value):
                btc_fees_cell.value = new_btc_value
                updates_btc += 1
            
            # Update USD fees cell if available
            if data_lookup[timestamp_str]['usd_fees'] is not None:
                usd_fees_cell = row[usd_fees_col_idx - 1]
                new_usd_value = data_lookup[timestamp_str]['usd_fees']
                
                if not math.isnan(new_usd_value):
                    usd_fees_cell.value = new_usd_value
                    updates_usd += 1
        
        # Progress indicator
        if rows_processed % 1000 == 0:
            print(f"Processed {rows_processed} rows...")
    
    # Save the updated workbook
    output_file = f"{excel_file.rsplit('.', 1)[0]}_manually_scaled.xlsx"
    workbook.save(output_file)
    print(f"Saved updated Excel file as {output_file}")
    
    # Print summary
    print("\nUpdate summary:")
    print(f"Total rows processed: {rows_processed}")
    print(f"Rows updated with transaction fees btc: {updates_btc}")
    print(f"Rows updated with Average transaction fees usd: {updates_usd}")

def main():
    # File to update
    excel_file = 'bitcoin_metrics_and_energy_updated_final.xlsx'
    
    # Step 1: Analyze the original fee data
    original_df = analyze_original_fees(excel_file)
    if original_df is None:
        print("Error: Could not analyze the original fee data")
        return
    
    # Step 2: Get the date range from the original data
    start_date = original_df['timestamp'].min().strftime('%Y-%m-%d')
    end_date = original_df['timestamp'].max().strftime('%Y-%m-%d')
    
    # Step 3: Fetch blockchain data for the same date range
    blockchain_df = fetch_and_analyze_blockchain_data(start_date, end_date)
    if blockchain_df.empty:
        print("Error: Could not fetch blockchain data")
        return
    
    # Step 4: Plot a comparison of the original and blockchain data
    plot_fee_comparison(original_df, blockchain_df)
    
    # Step 5: Manually scale the blockchain data
    scaled_df = manually_scale_blockchain_data(original_df, blockchain_df)
    
    # Step 6: Ensure 3-hour intervals through interpolation
    if scaled_df['timestamp'].diff().median().total_seconds() > 3.5 * 3600:
        print("Interpolating to ensure 3-hour intervals")
        scaled_df = interpolate_to_3_hour_intervals(scaled_df)
    
    # Step 7: Update the Excel file with the scaled data
    update_excel_with_scaled_fees(excel_file, scaled_df)

if __name__ == "__main__":
    main()
